#!/usr/bin/env python3
# Objective: preprocess for box014 only
# Tasbolat Taunyazov

import numpy as np
from pathlib import Path
import glob
from sklearn.model_selection import train_test_split
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook


grasp_data_dir = 'data/grasps_tight_lower'
save_dir = 'data/grasps_tight_lower/preprocessed2'
soft = False
TRANS_DISTANCE_MAX = 0.30 # in meters -> use this to filter out far grasps

test_info = {
    'mug': [2, 8, 14],
    'bottle': [3, 12, 19],
    'box':[14, 17],
    'bowl':[1, 16],
    'cylinder': [2, 11],
    'pan': [3,6],
    'scissor': [4,7],
    'fork': [1, 11],
    'hammer': [2, 15],
    'spatula': [1, 14]
}

exclude_info = {
    'box': [],
    'mug': [],
    'bowl': [5],
    'bottle': [],
    'cylinder': [],
}

excel_path = f'{grasp_data_dir}/info2.xlsx'


def save_info(info, cat):
    info = pd.DataFrame(info)

    if Path(excel_path).is_file():
        book = load_workbook(excel_path)
        writer = pd.ExcelWriter(excel_path, engine = 'openpyxl')
        if cat in book.sheetnames:
            _old_sheet = book[cat]
            book.remove(_old_sheet)
        writer.book = book
    else:
        writer = pd.ExcelWriter(excel_path, engine = 'openpyxl')

    info.to_excel(writer, sheet_name=cat)
    writer.save()

def get_filenames(cat='', idx=0):
    return glob.glob(f"{grasp_data_dir}/{cat}/{cat}{idx:03}_isaac*/*.npz")



def count_data(cat):
    df = None
    for i in range(0,21,1):
        data, df_info = read_data(cat, i, collect_info=True)

        if data is None:
            continue

        if df is None:
            df = df_info.copy()
        else:
            df = pd.concat([df, df_info])

        

    save_info(df, cat)

    del data
        

def read_data_soft(cat, i):

    # list all the files
    data = {}
    data['quaternions'] = []
    data['translations'] = []
    data['isaac_labels'] = []
    data['metadata']  = []

    # fnames = get_filenames(cat=cat, idx=i)
    # open main  grasp
    try:
        main_grasps = np.load(f"{grasp_data_dir}/{cat}/{cat}{i:03}_isaac/main_grasps.npz")
    except:
        return None

    for k in range(1000):

        try:
            sample_data = np.load(f"{grasp_data_dir}/{cat}/{cat}{i:03}_isaac/{k:08}.npz")
        except:
            return None

        data['isaac_labels'].append(np.sum(sample_data['isaac_labels'])/len(sample_data['isaac_labels']))

    data['isaac_labels'] = np.array(data['isaac_labels'])
    data['translations'] = main_grasps['translations']
    data['quaternions'] = main_grasps['quaternions']
    data['metadata'] = np.ones(len(data['translations']))*i

    avg_label_count = np.mean(data['isaac_labels'])
        
    print(f'{cat}{i:03}: {avg_label_count}')

    return data

def read_data(cat, i, collect_info=False):

    # list all the files
    data = {}
    data['quaternions'] = []
    data['translations'] = []
    data['isaac_labels'] = []
    data['metadata']  = []

    fnames = get_filenames(cat=cat, idx=i)
    if len(fnames) == 0:
        # raise ValueError('No files in the file')
        return None, None
    
    for fname in fnames:

        if 'main' in fname:
            continue

        sample_data = np.load(fname)

        grasp_trans_distance = np.linalg.norm(sample_data['translations'], axis=1)
        chosen_mask = (grasp_trans_distance <= TRANS_DISTANCE_MAX)

        if not('isaac_labels' in sample_data):
            continue

        data['isaac_labels'].append( sample_data['isaac_labels'][chosen_mask] )
        data['quaternions'].append( sample_data['quaternions'][chosen_mask] )
        data['translations'].append( sample_data['translations'][chosen_mask] )

    data['quaternions'] = np.concatenate(data['quaternions'])
    data['translations'] = np.concatenate(data['translations'])
    data['isaac_labels'] = np.concatenate(data['isaac_labels'])
    data['metadata'] = np.ones(len(data['translations']))*i

    pos_count = np.sum(data['isaac_labels'])
    neg_count = len(data['isaac_labels']) - pos_count
        
    print(f'{cat}{i:03}: {pos_count} {neg_count}')

    if pos_count == 0:
        return None, None

    if collect_info:
        info = {
        'cat': cat,
        'idx': [],
        'Pos': [],
        'Neg': [],
        'Total': [],
        'Pos_share_percent': [],
        'Ratio':[]
        }
        
        info['idx'].append(i)
        info['Pos'].append(pos_count)
        info['Neg'].append(neg_count)
        total = pos_count + neg_count
        info['Total'].append(total)
        _percentage = pos_count/total*100 if total!=0 else np.inf
        info['Pos_share_percent'].append(_percentage)       
        _ratio = neg_count/pos_count if pos_count !=0 else np.inf   
        info['Ratio'].append(_ratio)
        df_info = pd.DataFrame(info)
        return data, df_info    

    return data, None

def preprocess(cat, is_train=True):
    quaternions = []
    translations = []
    isaac_labels = []
    metadata = []
    for i in range(0,21,1):

        # if i in exclude_info[cat]:
        #     continue

        if i in test_info[cat]:
            if is_train:
                continue
        else:
            if not is_train:
                continue

        if soft:
            data = read_data_soft(cat=cat, i=i)
        else:
            data, _ = read_data(cat=cat, i=i)

        if data is None:
            continue

        quaternions.append( data['quaternions'] )
        translations.append( data['translations'] )
        isaac_labels.append( data['isaac_labels'] )
        metadata.append( data['metadata'] )
        del data

    quaternions = np.concatenate(quaternions)
    translations = np.concatenate(translations)
    isaac_labels = np.concatenate(isaac_labels)
    metadata = np.concatenate(metadata)

    # reorder from pos to neg

    print(f'Final count for {cat} is {len(isaac_labels)}:')
    print(f'Quaternions shape is {quaternions.shape}')
    print(f'Translations shape is {translations.shape}')
    print(f'isaac_labels shape is {isaac_labels.shape}')
    print()
    Path(f'{save_dir}/{cat}').mkdir(parents=True, exist_ok=True)

    if is_train:
        np.save(f'{save_dir}/{cat}/quaternions_train', quaternions)
        np.save(f'{save_dir}/{cat}/translations_train', translations)
        np.save(f'{save_dir}/{cat}/isaac_labels_train', isaac_labels)
        np.save(f'{save_dir}/{cat}/metadata_train', metadata)
        print('Done with train set preprocessing')
    else:
        np.save(f'{save_dir}/{cat}/quaternions_test', quaternions)
        np.save(f'{save_dir}/{cat}/translations_test', translations)
        np.save(f'{save_dir}/{cat}/isaac_labels_test', isaac_labels)
        np.save(f'{save_dir}/{cat}/metadata_test', metadata)
        print('Done with test set preprocessing')

    del quaternions, translations, isaac_labels

if __name__ == '__main__':
    cats = ['bowl', 'mug', 'bottle', 'box', 'cylinder', 'spatula', 'hammer', 'pan', 'fork', 'scissor']


    ######### FOR DATA PREPROCESS #################
    for cat in cats:
        preprocess(cat, is_train=False)
    for cat in cats:
        preprocess(cat, is_train=True)

    ##### FOR DATA COUNTING ##############
    # for cat in cats:
    #     count_data(cat)